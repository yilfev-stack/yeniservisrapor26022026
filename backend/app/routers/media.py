from datetime import datetime, timezone
from pathlib import Path

from fastapi import APIRouter, HTTPException, UploadFile
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from weasyprint import HTML

from app.db import collection
from app.schemas import ExcelExportOptionsIn, ExportOptionsIn
from app.storage import (
    EXPORT_DIR,
    UPLOAD_DIR,
    build_thumbnail_and_optimized,
    local_export_url,
    local_upload_url,
    save_original_image,
    upload_bytes_to_minio,
)
from .common import now, parse_id

router = APIRouter(prefix='/api', tags=['media'])


def _get_company_profile(report: dict):
    profile_id = report.get('company_profile_id')
    if profile_id:
        return profile_id
    return None


@router.post('/reports/{report_id}/photos')
async def upload_photo(report_id: str, kind: str, file: UploadFile, caption: str = '', tags: str = ''):
    if kind not in {'before', 'after'}:
        raise HTTPException(status_code=400, detail='kind must be before/after')

    raw = await file.read()
    original_rel, original_path = save_original_image(report_id, file.filename, raw)
    (
        optimized_rel,
        optimized_path,
        optimized_width,
        optimized_height,
        thumb_rel,
        thumb_path,
        thumb_width,
        thumb_height,
    ) = build_thumbnail_and_optimized(original_path, report_id, file.filename)

    upload_bytes_to_minio('demart-photos', original_rel, raw, file.content_type or 'image/jpeg')
    upload_bytes_to_minio('demart-photos', optimized_rel, optimized_path.read_bytes(), 'image/jpeg')
    upload_bytes_to_minio('demart-photos', thumb_rel, thumb_path.read_bytes(), 'image/jpeg')

    photo = {
        'report_id': report_id,
        'kind': kind,
        'caption': caption,
        'tags': [x.strip() for x in tags.split(',') if x.strip()],
        'original_object_key': original_rel,
        'original_size_bytes': len(raw),
        'optimized_object_key': optimized_rel,
        'thumb_object_key': thumb_rel,
        'optimized_width': optimized_width,
        'optimized_height': optimized_height,
        'thumb_width': thumb_width,
        'thumb_height': thumb_height,
        'created_at': datetime.now(timezone.utc),
    }
    inserted = await collection('photos').insert_one(photo)
    await collection('reports').update_one({'_id': parse_id(report_id)}, {'$push': {f'photo_sets.{kind}': str(inserted.inserted_id)}, '$set': {'updated_at': now()}})
    return {'id': str(inserted.inserted_id), 'thumb_url': local_upload_url(thumb_rel), 'optimized_url': local_upload_url(optimized_rel)}


@router.put('/photos/{photo_id}')
async def update_photo(photo_id: str, payload: dict):
    await collection('photos').update_one({'_id': parse_id(photo_id)}, {'$set': payload | {'updated_at': now()}})
    return {'ok': True}


@router.delete('/photos/{photo_id}')
async def delete_photo(photo_id: str):
    await collection('photos').delete_one({'_id': parse_id(photo_id)})
    return {'ok': True}


def _build_pdf_html(report: dict, before: list[dict], after: list[dict], options: ExportOptionsIn, company: dict | None):
    company_html = ''
    if company:
        company_html = f"""
        <div style='font-size:12px'>
          <strong>{company.get('name','')}</strong><br/>
          {company.get('address','')}<br/>
          {company.get('phone','')} · {company.get('email','')}
        </div>
        """

    slots = options.photos_per_page
    photo_width = '48%' if slots <= 4 else ('31%' if slots == 6 else '23%')

    def section(title: str, photos: list[dict]):
        blocks = []
        for p in photos:
            image_path = UPLOAD_DIR / p.get('optimized_object_key', '')
            if image_path.exists():
                blocks.append(
                    f"<div style='width:{photo_width}; margin:0 1% 14px 1%; display:inline-block; vertical-align:top;'>"
                    f"<img src='file://{image_path}' style='width:100%; height:170px; object-fit:contain; border:1px solid #ddd;'/>"
                    f"<div style='font-size:10px; color:#444; margin-top:4px'>{p.get('caption','')}</div></div>"
                )
        return f"<h3>{title}</h3><div>{''.join(blocks) or '<em>No photos</em>'}</div>"

    return f"""
    <html><body style='font-family: Arial, sans-serif; font-size:12px;'>
      <div style='border-bottom:2px solid #1e40af; padding-bottom:8px; margin-bottom:12px;'>
        <h1 style='margin:0; color:#1e40af'>SERVICE REPORT</h1>
        {company_html}
        <div>Report No: {report.get('report_no','')} | Revision: {report.get('revision_no',1)} | Language: {options.language}</div>
      </div>
      <h3>General</h3>
      <p>Customer: {report.get('customer_id','')} | Contact: {report.get('contact_id','')} | Status: {report.get('status','')}</p>
      <h3>Complaint</h3><p>{' '.join([x.get('text','') for x in report.get('blocks',{}).get('complaint',[])])}</p>
      <h3>Problems</h3><p>{' '.join([x.get('text','') for x in report.get('blocks',{}).get('problems',[])])}</p>
      <h3>Actions</h3><p>{' '.join([x.get('text','') for x in report.get('blocks',{}).get('actions',[])])}</p>
      <h3>Spares</h3><p>{', '.join([f"{x.get('part_name','')} x{x.get('qty','')}" for x in report.get('spares',[])])}</p>
      <h3>Result</h3><p>{report.get('result_notes','')}</p>
      {section('Before Photos', before)}
      {section('After Photos', after)}
    </body></html>
    """


@router.post('/reports/{report_id}/export/pdf')
async def export_pdf(report_id: str, payload: ExportOptionsIn):
    report = await collection('reports').find_one({'_id': parse_id(report_id)})
    if not report:
        raise HTTPException(status_code=404, detail='Report not found')

    photo_ids = (report.get('photo_sets') or {}).get('before', []) + (report.get('photo_sets') or {}).get('after', [])
    photos = [p async for p in collection('photos').find({'_id': {'$in': [parse_id(x) for x in photo_ids if x]}})] if photo_ids else []
    before = [p for p in photos if p.get('kind') == 'before']
    after = [p for p in photos if p.get('kind') == 'after']

    company = None
    profile_id = _get_company_profile(report)
    if profile_id:
        company = await collection('company_profiles').find_one({'_id': parse_id(profile_id)})
    if not company:
        company = await collection('company_profiles').find_one({'is_default': True})

    html = _build_pdf_html(report, before, after, payload, company)
    filename = f"{report.get('report_no', report_id)}-{payload.language}.pdf"
    file_path = EXPORT_DIR / filename
    HTML(string=html).write_pdf(file_path)

    export_doc = {
        'report_id': report_id,
        'type': 'pdf',
        'file_name': filename,
        'file_path': str(file_path),
        'options': payload.model_dump(),
        'created_at': now(),
    }
    inserted = await collection('exports').insert_one(export_doc)
    await collection('reports').update_one({'_id': report['_id']}, {'$set': {'exports.pdf': {'latest_url': local_export_url(filename), 'generated_at': now(), 'size_bytes': file_path.stat().st_size}}})
    return {'export_id': str(inserted.inserted_id), 'url': local_export_url(filename), 'size_bytes': file_path.stat().st_size}


@router.post('/reports/{report_id}/export/excel')
async def export_excel(report_id: str, payload: ExcelExportOptionsIn):
    report = await collection('reports').find_one({'_id': parse_id(report_id)})
    if not report:
        raise HTTPException(status_code=404, detail='Report not found')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Summary'
    ws['A1'] = 'Report No'
    ws['B1'] = report.get('report_no')
    ws['A2'] = 'Status'
    ws['B2'] = report.get('status')
    ws['A3'] = 'Result'
    ws['B3'] = report.get('result_notes', '')

    findings = wb.create_sheet('Findings')
    findings['A1'] = 'Problems'
    findings['A2'] = ' | '.join([x.get('text', '') for x in report.get('blocks', {}).get('problems', [])])

    actions = wb.create_sheet('Actions')
    actions['A1'] = 'Actions'
    actions['A2'] = ' | '.join([x.get('text', '') for x in report.get('blocks', {}).get('actions', [])])

    parts = wb.create_sheet('Parts')
    parts.append(['Part', 'Qty', 'Note'])
    for part in report.get('spares', []):
        parts.append([part.get('part_name'), part.get('qty'), part.get('note')])

    photos_ws = wb.create_sheet('Photos')
    photos_ws.append(['Before', 'Before Caption', 'After', 'After Caption'])
    before_ids = (report.get('photo_sets') or {}).get('before', [])
    after_ids = (report.get('photo_sets') or {}).get('after', [])
    before = [p async for p in collection('photos').find({'_id': {'$in': [parse_id(x) for x in before_ids if x]}})] if before_ids else []
    after = [p async for p in collection('photos').find({'_id': {'$in': [parse_id(x) for x in after_ids if x]}})] if after_ids else []

    rows = max(len(before), len(after))
    row_idx = 2
    for i in range(rows):
        b = before[i] if i < len(before) else None
        a = after[i] if i < len(after) else None
        photos_ws[f'B{row_idx}'] = b.get('caption', '') if b else ''
        photos_ws[f'D{row_idx}'] = a.get('caption', '') if a else ''
        if b:
            path = UPLOAD_DIR / b.get('optimized_object_key', '')
            if path.exists():
                img = XLImage(str(path))
                img.width, img.height = 180, 120
                photos_ws.add_image(img, f'A{row_idx}')
        if a:
            path = UPLOAD_DIR / a.get('optimized_object_key', '')
            if path.exists():
                img = XLImage(str(path))
                img.width, img.height = 180, 120
                photos_ws.add_image(img, f'C{row_idx}')
        photos_ws.row_dimensions[row_idx].height = 95
        row_idx += 6

    if payload.type == 'internal':
        wb.create_sheet('Measurements')
        wb.create_sheet('Work_Order')
        wb.create_sheet('History')

    filename = f"{report.get('report_no', report_id)}-{payload.type}-{payload.language}.xlsx"
    file_path = EXPORT_DIR / filename
    wb.save(file_path)

    export_doc = {
        'report_id': report_id,
        'type': f'excel_{payload.type}',
        'file_name': filename,
        'file_path': str(file_path),
        'options': payload.model_dump(),
        'created_at': now(),
    }
    inserted = await collection('exports').insert_one(export_doc)
    await collection('reports').update_one({'_id': report['_id']}, {'$set': {f'exports.excel_{payload.type}': {'latest_url': local_export_url(filename), 'generated_at': now(), 'size_bytes': file_path.stat().st_size}}})
    return {'export_id': str(inserted.inserted_id), 'url': local_export_url(filename), 'size_bytes': file_path.stat().st_size}


@router.get('/exports')
async def list_exports():
    return [
        {
            'id': str(doc['_id']),
            'type': doc.get('type'),
            'file_name': doc.get('file_name'),
            'url': local_export_url(doc.get('file_name')),
            'created_at': doc.get('created_at'),
        }
        async for doc in collection('exports').find().sort('created_at', -1)
    ]


@router.get('/exports/{export_id}/download')
async def download_export(export_id: str):
    doc = await collection('exports').find_one({'_id': parse_id(export_id)})
    if not doc:
        raise HTTPException(status_code=404, detail='Export not found')
    path = Path(doc['file_path'])
    if not path.exists():
        raise HTTPException(status_code=404, detail='File missing')
    return FileResponse(path)
